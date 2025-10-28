"""
Forward Test Results Export Module

Exports forward test results to Excel with:
- Performance metrics summary  
- All trades detailed listing
- Configuration dialog box text (exact copy from GUI)

SIMPLE RULE: GUI passes dialog_text, we paste it into Excel. That's it.
"""
import os
import pandas as pd
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, Any, List, Callable, Optional, Tuple, TYPE_CHECKING
import logging
from threading import Thread

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins
    
    # Import shared dashboard components
    from shared.dashboard_components import (
        DashboardStyleManager, 
        DashboardLayoutManager, 
        DashboardTableBuilder
    )
    
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.error("openpyxl is required for Excel export - no CSV fallback available")

# Type annotations for optional imports
if TYPE_CHECKING:
    from shared.dashboard_components import (
        DashboardStyleManager, 
        DashboardLayoutManager, 
        DashboardTableBuilder
    )

class ForwardTestResults:

    """Export forward test results to Excel"""

    

    def __init__(self, config: Dict[str, Any], position_manager, start_time: datetime, dialog_text: str = None):

        """

        Args:

            config: Configuration dictionary

            position_manager: Position manager with trade history

            start_time: Test start timestamp

            dialog_text: Dialog box text to paste into Excel (REQUIRED)

        """

        self.config = config

        self.position_manager = position_manager

        self.start_time = start_time

        self.end_time = None

        self.dialog_text = dialog_text

        

        # 🔍 DEBUG: Log what dialog_text we received

        logger.info(f"🔍 ForwardTestResults.__init__ - dialog_text type: {type(dialog_text)}, length: {len(dialog_text) if dialog_text else 0}")

        if dialog_text:

            logger.info(f"✅ Received dialog_text - first 100 chars: {dialog_text[:100]}")

        else:

            logger.warning(f"⚠️ dialog_text is empty or None: {repr(dialog_text)}")

    

    def finalize(self):

        """Mark the forward test as completed"""

        self.end_time = datetime.now()

    

    def export_to_excel(self, test_type: str = "live") -> str:

        """Export forward test results using shared dashboard components"""

        if not OPENPYXL_AVAILABLE:

            logger.error("Excel export not available - openpyxl required for forward test results")

            raise ImportError("openpyxl is required for forward test results export")

        

        # Automatic path detection: data simulation vs live webstream mode

        data_sim = self.config.get('data_simulation', {})

        if data_sim.get('enabled', False):

            detected_mode = "data"

            results_dir = r"C:\Users\user\Desktop\BotResults\results\Forward Test\data"

            logger.info("Path detection: Using data simulation mode")

        else:

            detected_mode = "live"

            results_dir = r"C:\Users\user\Desktop\BotResults\results\Forward Test\live"

            logger.info("Path detection: Using live webstream mode")

            

        # Create the specific directory

        os.makedirs(results_dir, exist_ok=True)

        

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Get symbol name for filename
        symbol = self.config.get('instrument', {}).get('symbol', 'UNKNOWN')
        symbol_clean = symbol.replace('/', '_').replace('\\', '_')  # Clean symbol for filename
        
        filename = os.path.join(results_dir, f"ft-{symbol_clean}-{timestamp}-{detected_mode}.xlsx")

        

        wb = Workbook()

        ws = wb.active

        ws.title = "Forward Test Results"

        

        # Validation: Check P&L consistency between summary and individual trades

        perf_summary = self.position_manager.get_performance_summary()

        trades_df = self._get_trades_dataframe()

        

        summary_net_pnl = perf_summary.get('total_pnl', 0)

        if not trades_df.empty and 'Net PnL' in trades_df.columns:

            trades_net_pnl_sum = trades_df['Net PnL'].fillna(0).sum()

            difference = abs(summary_net_pnl - trades_net_pnl_sum)

            

            logger.info(f"P&L Validation - Summary Net P&L: {summary_net_pnl:.2f}")

            logger.info(f"P&L Validation - Trades Net P&L Sum: {trades_net_pnl_sum:.2f}")

            logger.info(f"P&L Validation - Difference: {difference:.2f}")

            

            if difference > 0.01:  # Allow small floating point differences

                logger.warning(f"P&L mismatch detected: {difference:.2f} difference between summary and trades")

            else:

                logger.info("P&L validation passed: Summary and trades Net P&L match")

        else:

            logger.info("P&L Validation - No trades to validate")

        

        # Use new dashboard-based export
        self._create_dashboard_export(ws, detected_mode)
        
        # Save main dashboard sheet
        wb.save(filename)
        
        # --- HYBRID APPROACH: Add "Config" sheet with guaranteed text view ---
        logger.info("=" * 80)
        logger.info("ADDING CONFIG SHEET (CSV-STYLE TEXT VIEW)")
        logger.info("=" * 80)
        
        try:
            # Get dialog text and split into lines
            dialog_text = self._get_dialog_box_text()
            lines = dialog_text.splitlines()
            
            logger.info(f"📄 Config sheet: {len(lines)} lines of text")
            logger.info(f"📝 First line: {lines[0] if lines else 'EMPTY'}")
            
            # Build DataFrame with one column
            config_df = pd.DataFrame({"Configuration": lines})
            
            # Append new sheet to the same file using pandas ExcelWriter
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                config_df.to_excel(writer, sheet_name="Config", index=False)
                
                # Access the worksheet to format font size (2x increase) and bold
                worksheet = writer.sheets['Config']
                from openpyxl.styles import Font
                
                # Base font size 11pt → 22pt (2x) and make bold
                config_font = Font(name='Courier New', size=22, bold=True)
                header_font = Font(name='Courier New', size=22, bold=True)
                
                # Format header row (row 1)
                for cell in worksheet[1]:
                    cell.font = header_font
                
                # Format all data rows (row 2 onwards)
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = config_font
                
                # Set column width to accommodate larger text
                worksheet.column_dimensions['A'].width = 150  # Increased from default
            
            logger.info(f"✅ Config sheet added successfully with {len(lines)} rows")
            logger.info("✅ Config sheet font: 22pt (2x) bold for better readability")
            logger.info("✅ Traders now have both fancy dashboard AND guaranteed text view")
            
        except Exception as e:
            logger.error(f"⚠️ Failed to add Config sheet: {e}", exc_info=True)

        # --- SESSION LOG SHEET: Automatically add GUI Logs Tab content (3rd sheet) ---
        logger.info("=" * 80)
        logger.info("ADDING SESSION LOG SHEET (AUTOMATIC - NO USER INTERVENTION)")
        logger.info("=" * 80)
        
        try:
            # Get GUI log data if GUI instance is available
            gui_log_text = None
            if hasattr(self, 'gui_instance') and self.gui_instance:
                try:
                    gui_log_text = self.gui_instance._get_gui_log_text()
                    logger.info(f"📋 Retrieved GUI log text: {len(gui_log_text)} characters")
                except Exception as e:
                    logger.warning(f"Could not retrieve GUI log text: {e}")
            
            if gui_log_text and gui_log_text.strip():
                # Split log text into lines
                log_lines = gui_log_text.splitlines()
                
                logger.info(f"📋 Session log sheet: {len(log_lines)} lines of log text")
                logger.info(f"📝 First line: {log_lines[0] if log_lines else 'EMPTY'}")
                
                # Build DataFrame with one column for logs
                log_df = pd.DataFrame({"Session Log": log_lines})
                
                # Append log sheet to the same file using pandas ExcelWriter
                with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    log_df.to_excel(writer, sheet_name="SessionLog", index=False)
                    
                    # Access the worksheet to format font
                    worksheet = writer.sheets['SessionLog']
                    from openpyxl.styles import Font
                    
                    # Use readable font styling - 1.5x size (11pt → 17pt) and bold
                    log_font = Font(name='Courier New', size=17, bold=True)
                    header_font = Font(name='Courier New', size=17, bold=True)
                    
                    # Format header row (row 1)
                    for cell in worksheet[1]:
                        cell.font = header_font
                    
                    # Format all data rows (row 2 onwards)
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        for cell in row:
                            cell.font = log_font
                    
                    # Set column width to accommodate log text
                    worksheet.column_dimensions['A'].width = 120  # Wide enough for log lines
                
                logger.info(f"✅ Session log sheet added automatically with {len(log_lines)} log entries")
                logger.info("✅ GUI Logs tab content saved to 3rd worksheet (SessionLog)")
                logger.info("✅ Session log font: 17pt (1.5x) bold for better readability")
                
            else:
                logger.warning("⚠️ No GUI log text available - session log sheet skipped")
                
        except Exception as e:
            logger.error(f"⚠️ Failed to add Session Log sheet: {e}", exc_info=True)
            logger.warning("Main dashboard still saved - Config sheet is supplementary")
        
        # Enhanced logging with full path confirmation
        logger.info("=" * 80)
        logger.info("FORWARD TEST RESULTS EXPORT COMPLETED")
        logger.info("=" * 80)
        logger.info(f"Mode detected: {detected_mode.upper()}")
        logger.info(f"Results folder: {results_dir}")
        logger.info(f"Full file path: {filename}")
        logger.info(f"File size: {os.path.getsize(filename)} bytes")
        logger.info("File contains: Enhanced Excel report with dashboard + Config sheet")
        logger.info("=" * 80)
        
        print(f"✅ Enhanced Excel results saved: {filename}")
        print(f"📁 Results folder: {results_dir}")
        print(f"📋 Sheets: Dashboard + Config (guaranteed text view)")
        return filename

    

    def export_results(self, test_type: str = "live") -> Thread:

        """Export forward test results to Excel in background thread for GUI responsiveness"""

        worker = Thread(target=self._export_to_excel, args=(test_type,), daemon=True)

        worker.start()

        return worker

    

    def _export_to_excel(self, test_type: str):

        """Internal method containing the actual Excel export logic"""

        try:

            filename = self.export_to_excel(test_type)

            logger.info(f"Background Excel export completed: {filename}")

            return filename

        except Exception as e:

            logger.exception(f"Background Excel export failed: {e}")

            raise

    

    def _get_test_duration(self) -> str:

        """Get formatted test duration"""

        if not self.end_time:

            return "In Progress"

        

        # Strip tzinfo from aware datetimes to match naive datetimes before subtraction

        end_time_naive = self.end_time

        start_time_naive = self.start_time

        

        if hasattr(self.end_time, 'tzinfo') and self.end_time.tzinfo is not None:

            end_time_naive = self.end_time.replace(tzinfo=None)

        if hasattr(self.start_time, 'tzinfo') and self.start_time.tzinfo is not None:

            start_time_naive = self.start_time.replace(tzinfo=None)

            

        duration = end_time_naive - start_time_naive

        hours, remainder = divmod(int(duration.total_seconds()), 3600)

        minutes, seconds = divmod(remainder, 60)

        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    

    def _generate_config_text(self) -> str:

        """Generate configuration section as text"""

        lines = []

        

        # Instrument

        inst = self.config.get('instrument', {})

        lines.append(f"Symbol: {inst.get('symbol', 'N/A')}")

        lines.append(f"Exchange: {inst.get('exchange', 'N/A')}")

        lines.append(f"Product Type: {inst.get('product_type', 'N/A')}")

        lines.append(f"Lot Size: {inst.get('lot_size', 'N/A')}")

        lines.append("")

        

        # Capital & Risk

        capital = self.config.get('capital', {})

        risk = self.config.get('risk', {})

        lines.append(f"Initial Capital: {capital.get('initial_capital', 0):,.2f}")  # Removed currency symbol

        lines.append(f"Max Trades/Day: {risk.get('max_positions_per_day', 'N/A')}")

        lines.append(f"Risk per Trade: {risk.get('risk_per_trade_percent', 0)}%")

        lines.append("")

        

        # Data Source

        data_sim = self.config.get('data_simulation', {})

        if data_sim.get('enabled', False):

            lines.append(f"Data Source: File Simulation")

            lines.append(f"File: {data_sim.get('file_path', 'N/A')}")

        else:

            lines.append(f"Data Source: Live WebStream")

        lines.append("")

        

        return "\\n".join(lines)

    

    def _generate_performance_text(self) -> str:

        """Generate performance summary as text"""

        perf = self.position_manager.get_performance_summary()

        

        lines = []

        lines.append(f"Total Trades: {perf['total_trades']}")

        lines.append(f"Winning Trades: {perf['winning_trades']}")

        lines.append(f"Losing Trades: {perf['losing_trades']}")

        lines.append(f"Win Rate: {perf['win_rate']:.1f}%")

        lines.append(f"Total P&L: {perf['total_pnl']:,.2f}")  # Removed currency symbol for consistency

        lines.append(f"Profit Factor: {perf['profit_factor']:.2f}")

        lines.append("")

        

        # Capital performance  

        initial = self.position_manager.initial_capital

        current = self.position_manager.current_capital

        change = current - initial

        change_pct = (change / initial) * 100 if initial > 0 else 0

        

        lines.append(f"Initial Capital: {initial:,.2f}")      # Removed currency symbol for consistency

        lines.append(f"Final Capital: {current:,.2f}")       # Removed currency symbol for consistency

        lines.append(f"Capital Change: {change:,.2f} ({change_pct:+.2f}%)")  # Removed currency symbol

        

        return "\\n".join(lines)

    

    def _get_trades_dataframe(self) -> pd.DataFrame:

        """Get all trades as DataFrame with structural rigor matching backtest format"""

        if not self.position_manager.completed_trades:

            return pd.DataFrame()

            

        rows = []

        for i, trade in enumerate(self.position_manager.completed_trades, 1):

            # Ensure all numeric values are properly handled, converting None to 0

            entry_price = float(trade.entry_price) if trade.entry_price is not None else 0.0

            exit_price = float(trade.exit_price) if trade.exit_price is not None else 0.0

            quantity = int(trade.quantity) if trade.quantity is not None else 0

            gross_pnl = float(trade.gross_pnl) if trade.gross_pnl is not None else 0.0

            commission = float(trade.commission) if trade.commission is not None else 0.0

            net_pnl = float(trade.net_pnl) if trade.net_pnl is not None else 0.0

            duration_minutes = float(trade.duration_minutes) if trade.duration_minutes is not None else 0.0

            

            rows.append({

                '#': i,

                'Entry Time': trade.entry_time.strftime('%Y-%m-%d %H:%M:%S') if trade.entry_time else '',

                'Exit Time': trade.exit_time.strftime('%Y-%m-%d %H:%M:%S') if trade.exit_time else '',

                'Entry Price': round(entry_price, 2),

                'Exit Price': round(exit_price, 2),

                'Qty': quantity,

                'Gross PnL': round(gross_pnl, 2),

                'Commission': round(commission, 2),

                'Net PnL': round(net_pnl, 2),

                'Exit Reason': str(trade.exit_reason) if trade.exit_reason else '',

                'Duration (min)': round(duration_minutes, 2)

            })

            

        df = pd.DataFrame(rows)

        

        # Debug logging to help identify issues

        logger.info(f"Created trades DataFrame with {len(df)} rows and {len(df.columns)} columns")

        logger.info(f"DataFrame columns: {list(df.columns)}")

        if not df.empty:

            logger.info(f"Sample row data: {df.iloc[0].to_dict()}")

        

        return df

    

    def get_results_summary(self) -> Dict[str, Any]:

        """Get complete results summary for GUI display"""

        return {

            'config': self.config,

            'performance': self.position_manager.get_performance_summary(),

            'trades': self._get_trades_dataframe(),

            'duration': self._get_test_duration(),

            'start_time': self.start_time,

            'end_time': self.end_time

        }

    

    def get_summary_metrics(self) -> List[Tuple[str, str]]:

        """

        Prepare summary metrics with structural rigor matching backtest format.

        Returns list of (label, value) tuples with guaranteed P&L data consistency.

        """

        perf = self.position_manager.get_performance_summary()

        initial_capital = self.position_manager.initial_capital

        

        # Calculate net P&L from structured trade data for consistency

        trades_df = self._get_trades_dataframe()

        calculated_net_pnl = trades_df['Net PnL'].sum() if not trades_df.empty else 0.0

        

        # Metrics rearranged for better visibility and to prevent text truncation
        # Left Column (B→C): Winning Trades, Win Rate, Net P&L, Best Trade, Initial Capital, Return %
        # Right Column (I→H): Total Trades, Losing Trades, Gross P&L, Profit Factor, Worst Trade, Final Capital
        metrics = [

            ("Winning Trades", str(perf.get('winning_trades', 0))),   # Left (B→C)
            ("Total Trades", str(perf.get('total_trades', 0))),       # Right (I→H)

            ("Win Rate", f"{perf.get('win_rate', 0):.2f}%"),          # Left (B→C)
            ("Losing Trades", str(perf.get('losing_trades', 0))),     # Right (I→H)

            ("Net P&L", f"{calculated_net_pnl:,.2f}"),                # Left (B→C)
            ("Gross P&L", f"{perf.get('total_pnl', 0):,.2f}"),        # Right (I→H)

            ("Best Trade", f"{perf.get('max_win', 0):,.2f}"),         # Left (B→C)
            ("Profit Factor", f"{perf.get('profit_factor', 0):.2f}"), # Right (I→H)

            ("Initial Capital", f"{initial_capital:,.2f}"),           # Left (B→C)
            ("Worst Trade", f"{perf.get('max_loss', 0):,.2f}"),       # Right (I→H)

            ("Return %", f"{(calculated_net_pnl/initial_capital*100):,.2f}%" if initial_capital > 0 else "0.00%"),    # Left (B→C)
            ("Final Capital", f"{initial_capital + calculated_net_pnl:,.2f}")  # Right (I→H)

        ]

        

        return metrics

    

    def get_config_table(self) -> pd.DataFrame:

        """

        Dynamic configuration table - adapts to enabled/disabled features.

        Only includes parameters that are actually configured or enabled.

        """

        config_data = []

        

        if hasattr(self, 'config') and self.config:

            # Extract all sections from config

            inst = self.config.get('instrument', {})

            capital = self.config.get('capital', {})

            risk = self.config.get('risk', {})

            session = self.config.get('session', {})

            strategy = self.config.get('strategy', {})

            data_sim = self.config.get('data_simulation', {})

            

            # DATA SOURCE section - Always show

            self._add_data_source_params(config_data, data_sim)

            

            # INSTRUMENT & SESSION section - Always show core params

            self._add_instrument_params(config_data, inst)

            self._add_session_params(config_data, session)

            

            # RISK & CAPITAL MANAGEMENT section - Dynamic based on enabled features

            self._add_capital_params(config_data, capital)

            self._add_risk_params(config_data, risk)

            

            # STRATEGY & INDICATORS section - Only enabled indicators

            self._add_strategy_params(config_data, strategy)

            

            # DATA SOURCE DETAILS section - Technical info

            self._add_data_source_details(config_data, data_sim)

            

            # TEST EXECUTION INFO - Runtime data

            self._add_execution_params(config_data)

                

        return pd.DataFrame(config_data, columns=['Parameter', 'Value'])

    

    def _add_data_source_params(self, config_data: list, data_sim: dict):

        """Add data source parameters dynamically"""

        if data_sim.get('enabled', False):

            config_data.append(("Data Source Type", "FILE DATA SIMULATION"))

            if data_sim.get('file_path'):

                config_data.append(("Historical File", data_sim.get('file_path')))

            config_data.append(("Data Warning", "Uses HISTORICAL data, not live market prices"))

        else:

            config_data.append(("Data Source Type", "LIVE WEBSTREAM"))

            config_data.append(("Data Status", "Live market data"))

    

    def _add_instrument_params(self, config_data: list, inst: dict):

        """Add instrument parameters - always included"""

        config_data.append(("Symbol", inst.get('symbol', 'N/A')))

        config_data.append(("Exchange", inst.get('exchange', 'N/A')))

        config_data.append(("Product Type", inst.get('product_type', 'N/A')))

        

        if inst.get('lot_size') is not None:

            config_data.append(("Lot Size", inst.get('lot_size')))

        if inst.get('tick_size') is not None:

            config_data.append(("Tick Size", inst.get('tick_size')))

    

    def _add_session_params(self, config_data: list, session: dict):

        """Add session parameters dynamically"""

        # Session times

        if session.get('start_hour') is not None and session.get('start_min') is not None:

            config_data.append(("Session Start", f"{session.get('start_hour', 9):02d}:{session.get('start_min', 15):02d}"))

        if session.get('end_hour') is not None and session.get('end_min') is not None:

            config_data.append(("Session End", f"{session.get('end_hour', 15):02d}:{session.get('end_min', 30):02d}"))

        

        # Optional session features

        if session.get('auto_stop_enabled') is not None:

            config_data.append(("Auto Stop", "Enabled" if session.get('auto_stop_enabled') else "Disabled"))

        if session.get('max_loss_per_day') is not None:

            config_data.append(("Max Loss/Day", session.get('max_loss_per_day')))

    

    def _add_capital_params(self, config_data: list, capital: dict):

        """Add capital parameters dynamically"""

        if capital.get('initial_capital') is not None:

            config_data.append(("Initial Capital", capital.get('initial_capital')))

    

    def _add_risk_params(self, config_data: list, risk: dict):

        """Add risk management parameters - only if configured"""

        # Basic risk parameters

        if risk.get('max_positions_per_day') is not None:

            config_data.append(("Max Trades/Day", risk.get('max_positions_per_day')))

        if risk.get('base_sl_points') is not None:

            config_data.append(("Base Stop Loss", f"{risk.get('base_sl_points')} points"))

        

        # Take Profit - only if configured

        tp_points = risk.get('tp_points', [])

        tp_percents = risk.get('tp_percents', [])

        if tp_points and len(tp_points) > 0:

            config_data.append(("Take Profit Levels", f"{len(tp_points)} levels"))

            config_data.append(("TP Points", str(tp_points)))

            if tp_percents and len(tp_percents) > 0:

                config_data.append(("TP Percentages", str(tp_percents)))

        

        # Trail Stop - only if enabled

        if risk.get('use_trail_stop') is True:

            config_data.append(("Trail Stop", "Enabled"))

            if risk.get('trail_activation_points') is not None:

                config_data.append(("Trail Activation", f"{risk.get('trail_activation_points')} points"))

            if risk.get('trail_distance_points') is not None:

                config_data.append(("Trail Distance", f"{risk.get('trail_distance_points')} points"))

        elif risk.get('use_trail_stop') is False:

            config_data.append(("Trail Stop", "Disabled"))

        

        # Risk percentages

        if risk.get('risk_per_trade_percent') is not None:

            config_data.append(("Risk per Trade", f"{risk.get('risk_per_trade_percent')}%"))

        if risk.get('commission_percent') is not None:

            config_data.append(("Commission", f"{risk.get('commission_percent')}%"))

    

    def _add_strategy_params(self, config_data: list, strategy: dict):

        """Add strategy parameters - only enabled indicators and features"""

        # Strategy version (always include)

        config_data.append(("Strategy Version", "1"))

        

        # Green bars requirement

        if strategy.get('consecutive_green_bars') is not None and strategy.get('consecutive_green_bars') > 0:

            config_data.append(("Green Bars Required", strategy.get('consecutive_green_bars')))

        

        # Build dynamic enabled indicators summary

        enabled_indicators = []

        

        # EMA Crossover

        if strategy.get('use_ema_crossover') is True:

            fast_ema = strategy.get('fast_ema', 'N/A')

            slow_ema = strategy.get('slow_ema', 'N/A')

            enabled_indicators.append(f"EMA Crossover: Fast={fast_ema}, Slow={slow_ema}")

            config_data.append(("EMA Crossover Enabled", "True"))

            config_data.append(("Fast EMA", fast_ema))

            config_data.append(("Slow EMA", slow_ema))

        

        # MACD

        if strategy.get('use_macd') is True:

            enabled_indicators.append("MACD Signal")

            config_data.append(("MACD Enabled", "True"))

            # Add MACD specific parameters if they exist

            if strategy.get('macd_fast_period') is not None:

                config_data.append(("MACD Fast Period", strategy.get('macd_fast_period')))

            if strategy.get('macd_slow_period') is not None:

                config_data.append(("MACD Slow Period", strategy.get('macd_slow_period')))

            if strategy.get('macd_signal_period') is not None:

                config_data.append(("MACD Signal Period", strategy.get('macd_signal_period')))

        

        # RSI Filter

        if strategy.get('use_rsi_filter') is True:

            enabled_indicators.append("RSI Filter")

            config_data.append(("RSI Filter Enabled", "True"))

            if strategy.get('rsi_period') is not None:

                config_data.append(("RSI Period", strategy.get('rsi_period')))

            if strategy.get('rsi_overbought') is not None:

                config_data.append(("RSI Overbought", strategy.get('rsi_overbought')))

            if strategy.get('rsi_oversold') is not None:

                config_data.append(("RSI Oversold", strategy.get('rsi_oversold')))

        

        # VWAP

        if strategy.get('use_vwap') is True:

            enabled_indicators.append("VWAP")

            config_data.append(("VWAP Enabled", "True"))

            if strategy.get('vwap_period') is not None:

                config_data.append(("VWAP Period", strategy.get('vwap_period')))

        

        # Higher Timeframe Trend

        if strategy.get('use_htf_trend') is True:

            enabled_indicators.append("HTF Trend")

            config_data.append(("HTF Trend Enabled", "True"))

            if strategy.get('htf_timeframe') is not None:

                config_data.append(("HTF Timeframe", strategy.get('htf_timeframe')))

        

        # Consecutive Green Bars

        consecutive_bars = strategy.get('consecutive_green_bars', 0)

        if consecutive_bars > 0:

            enabled_indicators.append(f"Consecutive Green: {consecutive_bars} bars required")

        

        # Noise Filter (commonly used)

        if strategy.get('use_noise_filter') is True or len(enabled_indicators) > 0:

            noise_threshold = strategy.get('noise_threshold_percent', 0.01)

            enabled_indicators.append(f"Noise Filter: {noise_threshold}% threshold")

        

        # Add summary if any indicators are enabled

        if enabled_indicators:

            config_data.append(("Enabled Indicators", "; ".join(enabled_indicators)))

        else:

            config_data.append(("Enabled Indicators", "Basic Strategy Only"))

    

    def _add_data_source_details(self, config_data: list, data_sim: dict):

        """Add technical data source details"""

        if data_sim.get('enabled', False):

            config_data.append(("Mode", "File Simulation"))

            if data_sim.get('file_path'):

                config_data.append(("File Path", data_sim.get('file_path')))

            config_data.append(("Status", "Historical data replay"))

        else:

            config_data.append(("Mode", "Live WebStream"))

            config_data.append(("Status", "Live market connection"))

    

    def _add_execution_params(self, config_data: list):

        """Add test execution timing information"""

        if hasattr(self, 'start_time') and self.start_time:

            config_data.append(("Start Time", self.start_time.strftime("%Y-%m-%d %H:%M:%S")))

        if hasattr(self, 'end_time') and self.end_time:

            config_data.append(("End Time", self.end_time.strftime("%Y-%m-%d %H:%M:%S")))

    

    def _get_dialog_box_text(self) -> str:
        """
        Get the exact text that appears in the dialog box.
        Prioritizes self.dialog_text if provided (from GUI), otherwise generates from config.
        """
        # PRIORITY 1: Use dialog_text passed from GUI if available
        if hasattr(self, 'dialog_text') and self.dialog_text:
            logger.info(f"✅ Using dialog_text from GUI - length: {len(self.dialog_text)}")
            return self.dialog_text
        
        # PRIORITY 2: Generate from config as fallback
        logger.info("⚠️ No dialog_text from GUI, generating from config")
        if not hasattr(self, 'config') or not self.config:
            return "Configuration not available"
        
        lines = []
        lines.append("FORWARD TEST CONFIGURATION REVIEW")
        lines.append("=" * 80)
        lines.append("")
        
        # Data source section
        data_sim = self.config.get('data_simulation', {})
        if data_sim.get('enabled', False):
            lines.append("DATA SOURCE: 📁 FILE DATA SIMULATION")
            if data_sim.get('file_path'):
                lines.append(f"Historical file: {data_sim.get('file_path')}")
            lines.append("")
            lines.append("⚠️ This will use HISTORICAL data, not live market prices!")
        else:
            lines.append("DATA SOURCE: 🌐 LIVE WEBSTREAM")
            lines.append("✅ Live market data connection")
        lines.append("")
        
        # Instrument & Session
        inst = self.config.get('instrument', {})
        session = self.config.get('session', {})
        lines.append("INSTRUMENT & SESSION")
        lines.append("-" * 40)
        lines.append(f"Symbol:              {inst.get('symbol', 'N/A')}")
        lines.append(f"Exchange:            {inst.get('exchange', 'N/A')}")
        lines.append(f"Product Type:        {inst.get('product_type', 'N/A')}")
        lines.append(f"Lot Size:            {inst.get('lot_size', 'N/A')}")
        lines.append(f"Tick Size:           {inst.get('tick_size', 'N/A')}")
        if session.get('start_hour') is not None:
            lines.append(f"Session Start:       {session.get('start_hour'):02d}:{session.get('start_min', 0):02d}")
        if session.get('end_hour') is not None:
            lines.append(f"Session End:         {session.get('end_hour'):02d}:{session.get('end_min', 30):02d}")
        lines.append("")
        
        # Risk & Capital Management
        capital = self.config.get('capital', {})
        risk = self.config.get('risk', {})
        lines.append("RISK & CAPITAL MANAGEMENT")
        lines.append("-" * 40)
        if capital.get('initial_capital'):
            lines.append(f"Initial Capital:     {capital.get('initial_capital'):,.0f}")
        if risk.get('max_positions_per_day'):
            lines.append(f"Max Trades/Day:      {risk.get('max_positions_per_day')}")
        if risk.get('base_sl_points'):
            lines.append(f"Base Stop Loss:      {risk.get('base_sl_points')} points")
        lines.append("")
        
        # Strategy & Indicators
        strategy = self.config.get('strategy', {})
        lines.append("STRATEGY & INDICATORS")
        lines.append("-" * 40)
        lines.append("Strategy Version:    1")
        if strategy.get('consecutive_green_bars'):
            lines.append(f"Green Bars Required: {strategy.get('consecutive_green_bars')}")
        lines.append("")
        
        lines.append("Enabled Indicators:")
        if strategy.get('use_ema_crossover'):
            lines.append(f"  EMA Crossover:     Fast={strategy.get('fast_ema')}, Slow={strategy.get('slow_ema')}")
        if strategy.get('consecutive_green_bars', 0) > 0:
            lines.append(f"  Consecutive Green: {strategy.get('consecutive_green_bars')} bars required")
        lines.append("  Noise Filter:      0.01% threshold")
        lines.append("")
        
        lines.append("=" * 80)
        lines.append("Configuration auto-generated from settings")
        
        return "\n".join(lines)

    

    def get_trades_table(self) -> pd.DataFrame:

        """

        Prepare trades data as DataFrame with consistent column structure.

        Matches backtest trades table format and column order.

        """

        return self._get_trades_dataframe()

    

    def _create_dashboard_export(self, ws, detected_mode: str):

        """Create dashboard using shared components"""

        try:

            # Initialize dashboard managers

            style_manager = DashboardStyleManager(scale_factor=2.0)  # 2x font size increase

            layout_manager = DashboardLayoutManager(ws, max_columns=15, section_buffer=1, scale_factor=2.0)

            table_builder = DashboardTableBuilder(layout_manager, style_manager)

            

            # Create dashboard sections

            self._create_dashboard_sections(

                table_builder, detected_mode, style_manager, layout_manager

            )

            

            # Set page properties

            ws.page_margins = PageMargins(left=0.7, right=0.7, top=0.75, bottom=0.75)

            ws.print_options.horizontalCentered = True

            

        except Exception as e:

            logger.error(f"Dashboard export failed: {e}")

            raise

    

    def _create_dashboard_sections(

        self, 

        table_builder: "DashboardTableBuilder", 

        detected_mode: str,

        style_manager: "DashboardStyleManager",

        layout_manager: "DashboardLayoutManager"

    ):

        """Create all dashboard sections in correct order"""

        

        # 1. Title section

        table_builder.create_title_section(

            "FORWARD TEST RESULTS DASHBOARD",

            f"Mode: {detected_mode.upper()} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        )

        

        # 2. Key highlight metric (Net P&L)

        metrics = self.get_summary_metrics()

        net_pnl_value = 0.0

        net_pnl_display = "0.00"

        

        for label, value in metrics:

            if 'Net P&L' in label:

                try:

                    # Extract numeric value for comparison

                    net_pnl_value = float(str(value).replace(',', ''))

                    net_pnl_display = value

                    break

                except (ValueError, AttributeError):

                    pass

                    

        table_builder.create_highlight_metric(

            "NET P&L", net_pnl_display, net_pnl_value >= 0

        )

        

        # 3. Performance metrics table

        table_builder.create_metrics_table(metrics, "PERFORMANCE SUMMARY")

        

        # 4. Strategy Configuration - NOTE: Now in separate "Config" sheet

        # Manual cell merging in dashboard causes "zeros on border" issue

        # Full configuration is available in the pandas-generated "Config" sheet tab

        logger.info("ℹ️ Strategy Configuration section skipped in dashboard")

        logger.info("ℹ️ Full configuration available in separate 'Config' sheet")

        layout_manager.advance_row(1)  # Add spacing before trades table

        

        # 5. Detailed trades table

        trades_df = self.get_trades_table()

        if not trades_df.empty:

            table_builder.create_trades_table(trades_df, "DETAILED TRADES LOG")

        else:

            # Create empty trades message using table builder

            table_builder._create_no_trades_message("DETAILED TRADES LOG")
