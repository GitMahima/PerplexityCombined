"""
backtest/results.py - FIXED VERSION

Complete functionality with properly working Excel output.
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
from dataclasses import dataclass
from ..utils.time_utils import format_timestamp
import os

# Import openpyxl with proper error handling
try:
    from openpyxl import load_workbook, Workbook  # type: ignore[reportMissingModuleSource]
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side  # type: ignore[reportMissingModuleSource]
    from openpyxl.utils import get_column_letter  # type: ignore[reportMissingModuleSource]
    from openpyxl.worksheet.page import PageMargins  # type: ignore[reportMissingModuleSource]
    OPENPYXL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl not available. Excel functionality will be limited.")
    OPENPYXL_AVAILABLE = False
    # Define dummy classes for type hints
    class Workbook: pass
    class Font: pass
    class PatternFill: pass
    class Alignment: pass
    class Border: pass
    class Side: pass
    class PageMargins: pass
    def get_column_letter(n): return chr(64 + n)

import logging

# =====================================================
# HELPER FUNCTIONS AND DATA CLASSES
# =====================================================

def safe_divide(numerator, denominator, default=0.0):
    try:
        if denominator == 0:
            return default
        return numerator / denominator
    except:
        return default

def calculate_drawdown(equity_curve: List[float]) -> float:
    max_drawdown = 0.0
    peak = equity_curve[0] if equity_curve else 0
    for value in equity_curve:
        if value > peak:
            peak = value
        drawdown = (peak - value) / peak if peak > 0 else 0
        max_drawdown = max(max_drawdown, drawdown)
    return max_drawdown * 100

@dataclass
class TradeResult:
    """Single structured record for a completed trade."""
    entry_time: datetime
    exit_time: datetime
    entry_price: float
    exit_price: float
    quantity: int
    pnl: float
    commission: float
    exit_reason: str

@dataclass
class TradingMetrics:
    """Aggregate statistics for evaluation."""
    total_trades: int = 0
    winning_trades: int = 0
    losing_trades: int = 0
    win_rate: float = 0.0
    gross_profit: float = 0.0
    gross_loss: float = 0.0
    avg_win: float = 0.0
    avg_loss: float = 0.0
    total_pnl: float = 0.0
    total_commission: float = 0.0
    net_pnl: float = 0.0
    best_trade: float = 0.0
    worst_trade: float = 0.0
    return_percent: float = 0.0
    final_capital: float = 0.0
    profit_factor: float = 0.0
    drawdown_percent: float = 0.0

# =====================================================
# FIXED OPTIMIZATION CLASSES
# =====================================================

class StyleManager:
    """Centralized style management with larger, more readable fonts."""
    
    def __init__(self, scale_factor: float = 1.0):
        if not OPENPYXL_AVAILABLE:
            return
            
        self.scale_factor = scale_factor
        self._setup_fonts()
        self._setup_fills()
        self._setup_borders()
    
    def _setup_fonts(self):
        if not OPENPYXL_AVAILABLE:
            return
            
        base_size = int(12 * self.scale_factor)  # Increased base size from 10 to 12
        self.fonts = {
            'title': Font(size=base_size + 16, bold=True, color="FFFFFF"),     # 28
            'header': Font(size=base_size + 10, bold=True, color="FFFFFF"),    # 22
            'subheader': Font(size=base_size + 4, bold=True),                  # 16
            'normal': Font(size=base_size + 2),                                # 14
            'metric_label': Font(size=base_size + 2, bold=True),               # 14
            'metric_value': Font(size=base_size + 4, bold=True),               # 16
            'highlight': Font(size=base_size + 14, bold=True, color="FFFFFF"), # 26
            'trade_data': Font(size=base_size + 1)                             # 13 - Special for trade data
        }
    
    def _setup_fills(self):
        if not OPENPYXL_AVAILABLE:
            return
            
        self.fills = {
            'title': PatternFill(start_color="2E4BC6", end_color="2E4BC6", fill_type="solid"),
            'header': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            'summary': PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid"),
            'positive': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'negative': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'neutral': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        }
    
    def _setup_borders(self):
        if not OPENPYXL_AVAILABLE:
            return
            
        self.border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )

class LayoutManager:
    """Manages worksheet layout with proper spacing and responsive design."""
    
    def __init__(self, worksheet, max_columns: int = 15, buffer_size: int = 1):  # Increased columns
        if not OPENPYXL_AVAILABLE:
            return
            
        self.ws = worksheet
        self.max_columns = max_columns
        self.buffer_size = buffer_size
        self.current_row = 1 + buffer_size
        self.setup_buffers()
    
    def setup_buffers(self):
        """Set up buffer columns and basic page setup."""
        if not OPENPYXL_AVAILABLE:
            return
            
        # Left buffer
        for i in range(self.buffer_size):
            col_letter = get_column_letter(i + 1)
            self.ws.column_dimensions[col_letter].width = 4
        
        # Page margins
        self.ws.page_margins = PageMargins(left=1.0, right=0.7, top=0.75, bottom=0.75)
    
    def get_usable_columns(self) -> Tuple[int, int]:
        """Return start and end column indices for content."""
        start_col = self.buffer_size + 1  # Start at column B (2)
        end_col = self.buffer_size + self.max_columns  # End at column P (16) for 15 columns
        return start_col, end_col
    
    def advance_row(self, rows: int = 1, add_spacing: bool = False):
        """Advance current row position."""
        self.current_row += rows
        if add_spacing:
            self.current_row += 1
    
    def merge_and_style_range(self, start_col: int, end_col: int, rows: int = 1, 
                            value: str = "", style_type: str = "normal") -> Any:
        """Merge cells in range and apply styling."""
        if not OPENPYXL_AVAILABLE:
            return None
            
        start_cell = f"{get_column_letter(start_col)}{self.current_row}"
        end_cell = f"{get_column_letter(end_col)}{self.current_row + rows - 1}"
        
        if start_col != end_col or rows > 1:
            self.ws.merge_cells(f"{start_cell}:{end_cell}")
        
        cell = self.ws[start_cell]
        cell.value = value
        return cell

class TableBuilder:
    """Builds different types of tables with consistent formatting and WORKING trade data."""
    
    def __init__(self, layout_manager: LayoutManager, style_manager: StyleManager):
        self.layout = layout_manager
        self.style = style_manager
    
    def create_title_section(self, title: str, subtitle: str = None):
        """Create main title section."""
        if not OPENPYXL_AVAILABLE:
            return
            
        start_col, end_col = self.layout.get_usable_columns()
        
        # Main title
        title_cell = self.layout.merge_and_style_range(start_col, end_col, 1, title)
        title_cell.font = self.style.fonts['title']
        title_cell.fill = self.style.fills['title']
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.layout.ws.row_dimensions[self.layout.current_row].height = 55
        self.layout.advance_row(1, add_spacing=True)
        
        # Subtitle if provided
        if subtitle:
            subtitle_cell = self.layout.merge_and_style_range(start_col, end_col, 1, subtitle)
            subtitle_cell.font = self.style.fonts['header']
            subtitle_cell.fill = self.style.fills['header']
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.layout.advance_row(1, add_spacing=True)
    
    def create_highlight_metric(self, label: str, value: str, is_positive: bool = None):
        """Create a highlighted key metric display."""
        if not OPENPYXL_AVAILABLE:
            return
            
        start_col, end_col = self.layout.get_usable_columns()
        center_start = start_col + 3  # Column E
        center_end = end_col - 3      # Column M (for 15 total columns B-P)
        
        # Label
        label_cell = self.layout.merge_and_style_range(center_start, center_end, 1, label)
        label_cell.font = Font(size=20, bold=True)  # Larger label font
        label_cell.fill = self.style.fills['summary']
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = self.style.border
        self.layout.ws.row_dimensions[self.layout.current_row].height = 65
        self.layout.advance_row(1)
        
        # Value
        value_cell = self.layout.merge_and_style_range(center_start, center_end, 1, value)
        value_cell.font = self.style.fonts['highlight']
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = self.style.border
        
        if is_positive is not None:
            value_cell.fill = self.style.fills['positive' if is_positive else 'negative']
        else:
            value_cell.fill = self.style.fills['title']
        
        self.layout.ws.row_dimensions[self.layout.current_row].height = 85
        self.layout.advance_row(1, add_spacing=True)
    
    def create_metrics_table(self, metrics_data: List[Tuple[str, Any]], title: str = "PERFORMANCE SUMMARY"):
        """Create a responsive metrics table that accommodates all data."""
        if not OPENPYXL_AVAILABLE:
            return
            
        start_col, end_col = self.layout.get_usable_columns()
        
        # Title
        title_cell = self.layout.merge_and_style_range(start_col, end_col, 1, title)
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.layout.ws.row_dimensions[self.layout.current_row].height = 45
        self.layout.advance_row(1)
        
        # Write metrics in two-column pairs with fixed positioning for better visibility
        pairs_per_row = 2  # Fixed: Left and Right pairs
        
        for i in range(0, len(metrics_data), pairs_per_row):
            row_pairs = metrics_data[i:i + pairs_per_row]
            
            for j, (label, value) in enumerate(row_pairs):
                # Fixed column positions for better visibility (per user requirements)
                # Left side: Label in C(3), Value in D(4)  
                # Right side: Label in H(8), Value in I(9)
                if j == 0:  # Left side metrics
                    label_col = 3  # Column C 
                    value_col = 4  # Column D
                else:  # Right side metrics
                    label_col = 8  # Column H (moved from I)
                    value_col = 9  # Column I (moved from L)
                
                # Label
                label_cell = self.layout.ws.cell(row=self.layout.current_row, column=label_col, value=label)
                label_cell.font = self.style.fonts['metric_label']
                label_cell.border = self.style.border
                label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Value
                value_cell = self.layout.ws.cell(row=self.layout.current_row, column=value_col, value=value)
                value_cell.font = self.style.fonts['metric_value']
                value_cell.border = self.style.border
                value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            self.layout.ws.row_dimensions[self.layout.current_row].height = 45  # Increased height
            self.layout.advance_row(1)
        
        self.layout.advance_row(0, add_spacing=True)
    
    def create_config_table(self, config_data: pd.DataFrame, title: str = "STRATEGY CONFIGURATION"):
        """Create configuration table with proper formatting."""
        if not OPENPYXL_AVAILABLE:
            return
            
        start_col, end_col = self.layout.get_usable_columns()
        
        # Title
        title_cell = self.layout.merge_and_style_range(start_col, end_col, 1, title)
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.layout.ws.row_dimensions[self.layout.current_row].height = 50
        self.layout.advance_row(1)
        
        # Configuration rows
        for _, row in config_data.iterrows():
            # Parameter name (4 columns)
            param_start = start_col
            param_end = start_col + 3
            param_cell = self.layout.merge_and_style_range(param_start, param_end, 1, row['Key'])
            param_cell.font = self.style.fonts['subheader']
            param_cell.fill = self.style.fills['summary']
            param_cell.border = self.style.border
            param_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Parameter value (remaining columns)
            value_start = start_col + 4
            value_end = end_col
            value_cell = self.layout.merge_and_style_range(value_start, value_end, 1)
            
            # Format multi-part values with line breaks
            value_text = str(row['Value'])
            if ',' in value_text and any(k in value_text.lower() for k in ['ema', 'fast', 'slow', 'enabled', 'activation']):
                parts = [p.strip() for p in value_text.split(',')]
                formatted_value = '\n'.join(f"• {p}" for p in parts)
                line_count = len(parts)
                self.layout.ws.row_dimensions[self.layout.current_row].height = line_count * 20
            else:
                formatted_value = value_text
                self.layout.ws.row_dimensions[self.layout.current_row].height = 30
            
            value_cell.value = formatted_value
            value_cell.font = self.style.fonts['normal']
            value_cell.border = self.style.border
            value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            self.layout.advance_row(1)
        
        self.layout.advance_row(0, add_spacing=True)
    
    def create_trades_table(self, trades_data: pd.DataFrame, title: str = "DETAILED TRADES LOG"):
        """Create WORKING trades table with actual data and all columns."""
        if not OPENPYXL_AVAILABLE or trades_data.empty:
            return
        
        start_col, end_col = self.layout.get_usable_columns()
        
        # Title
        title_cell = self.layout.merge_and_style_range(start_col, end_col, 1, title)
        title_cell.font = self.style.fonts['header']
        title_cell.fill = self.style.fills['header']
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        self.layout.ws.row_dimensions[self.layout.current_row].height = 30
        self.layout.advance_row(1)
        
        # Define ALL columns we want to show
        all_columns = [
            '#', 'Entry Time', 'Exit Time', 'Entry ₹', 'Exit ₹', 
            'Lots', 'Total Qty', 'Gross P&L', 'Commission', 'Net P&L', 
            'Exit Reason', 'Duration (min)', 'Capital Outstanding'
        ]
        
        # Calculate how many columns we can fit
        available_columns = end_col - start_col + 1
        
        if len(all_columns) <= available_columns:
            # Single table - all columns fit
            self._create_working_trades_table(trades_data, start_col, all_columns)
        else:
            # Split into two tables
            mid_point = 7  # Split after Net P&L column
            table1_columns = all_columns[:mid_point]
            table2_columns = ['#'] + all_columns[mid_point:]  # Include # for reference
            
            # First table
            self._create_working_trades_table(trades_data, start_col, table1_columns)
            
            # Spacing
            self.layout.advance_row(2, add_spacing=True)
            
            # Second table title
            title2_cell = self.layout.merge_and_style_range(start_col, end_col, 1, f"{title} (Continued)")
            title2_cell.font = self.style.fonts['header']
            title2_cell.fill = self.style.fills['header']
            title2_cell.alignment = Alignment(horizontal="center", vertical="center")
            self.layout.advance_row(1)
            
            # Second table
            self._create_working_trades_table(trades_data, start_col, table2_columns)
    
    def _create_working_trades_table(self, trades_data: pd.DataFrame, start_col: int, columns: List[str]):
        """Create a single working trades table with actual data."""
        if not OPENPYXL_AVAILABLE:
            return
            
        # Headers
        for col_idx, header in enumerate(columns):
            if start_col + col_idx > 16:  # Don't exceed our column limit
                break
            header_cell = self.layout.ws.cell(
                row=self.layout.current_row, 
                column=start_col + col_idx, 
                value=header
            )
            header_cell.font = self.style.fonts['header']
            header_cell.fill = self.style.fills['header']
            header_cell.border = self.style.border
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.layout.advance_row(1)
        
        # Data rows with ACTUAL trade data
        for trade_idx, (_, trade_row) in enumerate(trades_data.iterrows(), 1):
            
            # Prepare the actual data for this row
            row_data = []
            for col in columns:
                if col == '#':
                    row_data.append(trade_idx)
                elif col == 'Entry Time':
                    if pd.notna(trade_row.get('Entry Time')):
                        row_data.append(pd.to_datetime(trade_row['Entry Time']).strftime('%m/%d %H:%M'))
                    else:
                        row_data.append('')
                elif col == 'Exit Time':
                    if pd.notna(trade_row.get('Exit Time')):
                        row_data.append(pd.to_datetime(trade_row['Exit Time']).strftime('%m/%d %H:%M'))
                    else:
                        row_data.append('')
                elif col == 'Entry ₹':
                    if pd.notna(trade_row.get('Entry Price')):
                        row_data.append(f"{trade_row['Entry Price']:.2f}")
                    else:
                        row_data.append('')
                elif col == 'Exit ₹':
                    if pd.notna(trade_row.get('Exit Price')):
                        row_data.append(f"{trade_row['Exit Price']:.2f}")
                    else:
                        row_data.append('')
                elif col == 'Lots':
                    row_data.append(trade_row.get('Lots', 'N/A'))
                elif col == 'Total Qty':
                    if pd.notna(trade_row.get('Total Qty')):
                        row_data.append(int(trade_row['Total Qty']))
                    else:
                        row_data.append('')
                elif col == 'Gross P&L':
                    if pd.notna(trade_row.get('Gross P&L')):
                        row_data.append(f"₹{trade_row['Gross P&L']:,.0f}")
                    else:
                        row_data.append('')
                elif col == 'Commission':
                    if pd.notna(trade_row.get('Commission')):
                        row_data.append(f"₹{trade_row['Commission']:,.0f}")
                    else:
                        row_data.append('')
                elif col == 'Net P&L':
                    if pd.notna(trade_row.get('Net P&L')):
                        row_data.append(f"₹{trade_row['Net P&L']:,.0f}")
                    else:
                        row_data.append('')
                elif col == 'Exit Reason':
                    reason = str(trade_row.get('Exit Reason', ''))
                    row_data.append(reason[:20] if len(reason) > 20 else reason)  # Limit length
                elif col == 'Duration (min)':
                    if pd.notna(trade_row.get('Duration (min)')):
                        row_data.append(f"{trade_row['Duration (min)']:.1f}")
                    else:
                        row_data.append('')
                elif col == 'Capital Outstanding':
                    if pd.notna(trade_row.get('Capital Outstanding')):
                        row_data.append(f"₹{trade_row['Capital Outstanding']:,.0f}")
                    else:
                        row_data.append('')
                else:
                    row_data.append('')
            
            # Write the actual data to cells
            for col_idx, value in enumerate(row_data):
                if start_col + col_idx > 16:  # Don't exceed our column limit
                    break
                    
                cell = self.layout.ws.cell(
                    row=self.layout.current_row,
                    column=start_col + col_idx,
                    value=value
                )
                cell.font = self.style.fonts['trade_data']  # Use trade-specific font
                cell.border = self.style.border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Color code P&L columns
                if columns[col_idx] in ['Gross P&L', 'Net P&L'] and isinstance(value, str):
                    if '₹' in value:
                        try:
                            # Extract numeric value
                            numeric_part = value.replace('₹', '').replace(',', '').strip()
                            if numeric_part and numeric_part != '':
                                numeric_val = float(numeric_part)
                                if numeric_val > 0:
                                    cell.fill = self.style.fills['positive']
                                elif numeric_val < 0:
                                    cell.fill = self.style.fills['negative']
                        except ValueError:
                            pass
            
            # Set row height
            self.layout.ws.row_dimensions[self.layout.current_row].height = 35
            self.layout.advance_row(1)
        
        # Set column widths for better readability
        self._set_working_column_widths(start_col, len(columns))
    
    def _set_working_column_widths(self, start_col: int, num_columns: int):
        """Set optimal column widths for trades table with larger sizes."""
        if not OPENPYXL_AVAILABLE:
            return
            
        # Larger column widths for better readability
        widths = [8, 18, 18, 14, 14, 10, 12, 16, 14, 16, 22, 14, 18]
        
        for i in range(num_columns):
            if start_col + i <= 16:
                col_letter = get_column_letter(start_col + i)
                width = widths[i] if i < len(widths) else 14
                self.layout.ws.column_dimensions[col_letter].width = width

class DataPreparator:
    """Prepares data for presentation in various formats."""
    
    def __init__(self, results_instance):
        self.results = results_instance
        self.metrics = results_instance.calculate_metrics()
    
    def get_metrics_data(self) -> List[Tuple[str, str]]:
        """Get formatted metrics as list of (label, value) tuples."""
        return [
            ("Total Trades", str(self.metrics.total_trades)),
            ("Win Rate", f"{self.metrics.win_rate:.2f}%"),
            ("Winning Trades", str(self.metrics.winning_trades)),
            ("Losing Trades", str(self.metrics.losing_trades)),
            ("Gross P&L", f"₹{self.metrics.total_pnl:,.2f}"),
            ("Commission", f"₹{self.metrics.total_commission:,.2f}"),
            ("Net P&L", f"₹{self.metrics.net_pnl:,.2f}"),
            ("Return %", f"{self.metrics.return_percent:.2f}%"),
            ("Best Trade", f"₹{self.metrics.best_trade:,.2f}"),
            ("Worst Trade", f"₹{self.metrics.worst_trade:,.2f}"),
            ("Avg Win", f"₹{self.metrics.avg_win:,.2f}"),
            ("Avg Loss", f"₹{self.metrics.avg_loss:,.2f}"),
            ("Start Capital", f"₹{self.results.initial_capital:,.2f}"),
            ("Final Capital", f"₹{self.metrics.final_capital:,.2f}"),
            ("Profit Factor", f"{self.metrics.profit_factor:.2f}"),
            ("Drawdown", f"{self.metrics.drawdown_percent:.2f}%")
        ]

# =====================================================
# MAIN RESULTS CLASS (COMPLETE & WORKING)
# =====================================================

class Results:
    """
    Result engine. Stores completed trades, tracks equity progression,
    and outputs performance metrics and reports with optimized Excel output.
    """

    def __init__(self, initial_capital: float):
        self.initial_capital = initial_capital
        self.current_capital = initial_capital
        self.trades: List[TradeResult] = []
        self.equity_curve: List[Tuple[datetime, float]] = []
        self.config: Optional[Dict[str, Any]] = None

    def add_trade(self, trade_data: Dict[str, Any]) -> None:
        """Appends a trade and updates capital/equity."""
        trade = TradeResult(
            entry_time=trade_data['entry_time'],
            exit_time=trade_data['exit_time'],
            entry_price=trade_data['entry_price'],
            exit_price=trade_data['exit_price'],
            quantity=trade_data['quantity'],
            pnl=trade_data['pnl'],
            commission=trade_data['commission'],
            exit_reason=trade_data.get('exit_reason', ''),
        )

        self.trades.append(trade)
        self.current_capital += (trade.pnl - trade.commission)
        self.equity_curve.append((trade.exit_time, self.current_capital))

    def set_config(self, config: Dict[str, Any]):
        self.config = config

    def _create_additional_info_table(self) -> pd.DataFrame:
        """Create additional info table with indicators and parameters"""
        if not hasattr(self, 'config') or not self.config:
            return pd.DataFrame([{"Key": "Configuration", "Value": "Not Available"}])
        
        rows = []
        config = self.config
        strategy_config = config.get('strategy', config)  # fallback if not nested
        risk_config = config.get('risk', config)  # fallback if not nested

        # Indicators activated
        indicator_map = {
            'use_ema_crossover': 'EMA Crossover',
            'use_macd': 'MACD',
            'use_vwap': 'VWAP',
            'use_rsi_filter': 'RSI Filter',
            'use_htf_trend': 'HTF Trend',
            'use_bollinger_bands': 'Bollinger Bands',
            'use_stochastic': 'Stochastic',
            'use_atr': 'ATR'
        }

        active_indicators = [name for key, name in indicator_map.items()
                           if strategy_config.get(key, False)]
        rows.append({"Key": "Indicators Activated",
                    "Value": ", ".join(active_indicators) if active_indicators else "None"})

        # EMA parameters
        if strategy_config.get('use_ema_crossover', False):
            ema_params = f"Fast EMA: {strategy_config.get('fast_ema', 9)}, Slow EMA: {strategy_config.get('slow_ema', 21)}"
            rows.append({"Key": "EMA Parameters", "Value": ema_params})

        # MACD parameters
        if strategy_config.get('use_macd', False):
            macd_params = f"Fast: {strategy_config.get('macd_fast', 12)}, Slow: {strategy_config.get('macd_slow', 26)}, Signal: {strategy_config.get('macd_signal', 9)}"
            rows.append({"Key": "MACD Parameters", "Value": macd_params})

        # RSI parameters
        if strategy_config.get('use_rsi_filter', False):
            rsi_params = f"Period: {strategy_config.get('rsi_length', 14)}, Overbought: {strategy_config.get('rsi_overbought', 70)}, Oversold: {strategy_config.get('rsi_oversold', 30)}"
            rows.append({"Key": "RSI Parameters", "Value": rsi_params})

        # HTF Trend parameters
        if strategy_config.get('use_htf_trend', False):
            htf_params = f"HTF Period: {strategy_config.get('htf_period', 20)}"
            rows.append({"Key": "HTF Trend Parameters", "Value": htf_params})

        # SL points
        rows.append({"Key": "SL Points", "Value": str(risk_config.get('base_sl_points', 15))})

        # TP points
        tp_points = risk_config.get('tp_points', config.get('tp_points', [10, 25, 50, 100]))
        rows.append({"Key": "TP Points", "Value": ", ".join(map(str, tp_points))})

        # Trail SL info
        trail_enabled = risk_config.get('use_trail_stop', config.get('use_trail_stop', True))
        trail_activation = risk_config.get('trail_activation_points', config.get('trail_activation_points', 25))
        trail_distance = risk_config.get('trail_distance_points', config.get('trail_distance_points', 10))
        trail_info = f"Enabled: {trail_enabled}, Activation: {trail_activation} points, Distance: {trail_distance} points"
        rows.append({"Key": "Trailing Stop", "Value": trail_info})

        # Green bars requirement
        green_bars_req = strategy_config.get('consecutive_green_bars')
        rows.append({"Key": "Green Bars Required for Entry", "Value": str(green_bars_req)})

        return pd.DataFrame(rows)

    def calculate_metrics(self) -> TradingMetrics:
        if not self.trades:
            return TradingMetrics(final_capital=self.initial_capital)

        total_trades = len(self.trades)
        wins = [t for t in self.trades if t.pnl > 0]
        losses = [t for t in self.trades if t.pnl < 0]

        gross_profit = sum(t.pnl for t in wins)
        gross_loss = sum(t.pnl for t in losses)
        total_commission = sum(t.commission for t in self.trades)
        total_pnl = gross_profit + gross_loss
        net_pnl = total_pnl - total_commission

        avg_win = safe_divide(gross_profit, len(wins))
        avg_loss = safe_divide(abs(gross_loss), len(losses))
        final_capital = self.initial_capital + net_pnl
        return_percent = safe_divide(net_pnl, self.initial_capital, 0.0) * 100
        drawdown = calculate_drawdown([v for _, v in self.equity_curve])

        return TradingMetrics(
            total_trades=total_trades,
            winning_trades=len(wins),
            losing_trades=len(losses),
            win_rate=100 * safe_divide(len(wins), total_trades),
            gross_profit=gross_profit,
            gross_loss=gross_loss,
            avg_win=avg_win,
            avg_loss=avg_loss,
            total_pnl=total_pnl,
            total_commission=total_commission,
            net_pnl=net_pnl,
            best_trade=max((t.pnl for t in self.trades), default=0.0),
            worst_trade=min((t.pnl for t in self.trades), default=0.0),
            return_percent=return_percent,
            final_capital=final_capital,
            profit_factor=safe_divide(gross_profit, abs(gross_loss), 0.0),
            drawdown_percent=drawdown,
        )

    def print_summary(self):
        """Show performance metrics in console."""
        m = self.calculate_metrics()
        print(f"\n{'='*60}")
        print("TRADING PERFORMANCE SUMMARY")
        print(f"{'='*60}")
        print(f"Total Trades     : {m.total_trades}")
        print(f"Win Rate (%)     : {m.win_rate:.2f}")
        print(f"Gross Profit     : ₹{m.gross_profit:.2f}")
        print(f"Gross Loss       : ₹{m.gross_loss:.2f}")
        print(f"Avg Win          : ₹{m.avg_win:.2f}")
        print(f"Avg Loss         : ₹{m.avg_loss:.2f}")
        print(f"Net P&L          : ₹{m.net_pnl:.2f}")
        print(f"Best Trade (P&L) : ₹{m.best_trade:.2f}")
        print(f"Worst Trade (P&L): ₹{m.worst_trade:.2f}")
        print(f"Return (%)       : {m.return_percent:.2f}")
        print(f"Drawdown (%)     : {m.drawdown_percent:.2f}")
        print(f"Final Capital    : ₹{m.final_capital:,.2f}")
        print(f"Profit Factor    : {m.profit_factor:.2f}")
        print(f"Total Commission : ₹{m.total_commission:.2f}")
        print(f"{'='*60}")

    def get_trade_summary(self) -> pd.DataFrame:
        rows = []
        capital = self.initial_capital

        # Insert starting capital row
        rows.append({
            "Entry Time": "",
            "Exit Time": "",
            "Entry Price": "",
            "Exit Price": "",
            "Lots": "",
            "Total Qty": "",
            "Gross P&L": "",
            "Commission": "",
            "Net P&L": "",
            "Exit Reason": "Starting Capital",
            "Duration (min)": "",
            "Capital Outstanding": round(capital, 2)
        })

        for t in self.trades:
            net_pnl = t.pnl - t.commission
            capital += net_pnl
            lots_display = getattr(t, 'lots_traded', t.quantity // getattr(t, 'lot_size', 1)) if hasattr(t, 'lot_size') else 'N/A'

            rows.append({
                "Entry Time": t.entry_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Exit Time": t.exit_time.strftime("%Y-%m-%d %H:%M:%S"),
                "Entry Price": round(t.entry_price, 2),
                "Exit Price": round(t.exit_price, 2),
                "Lots": lots_display,
                "Total Qty": t.quantity,
                "Gross P&L": round(t.pnl, 2),
                "Commission": round(t.commission, 2),
                "Net P&L": round(net_pnl, 2),
                "Exit Reason": t.exit_reason,
                "Duration (min)": round((t.exit_time - t.entry_time).total_seconds() / 60, 2),
                "Capital Outstanding": round(capital, 2)
            })

        return pd.DataFrame(rows)

    def get_equity_curve(self) -> pd.DataFrame:
        """Return equity curve as DataFrame."""
        if not self.equity_curve:
            return pd.DataFrame()
        return pd.DataFrame(self.equity_curve, columns=["timestamp", "capital"])

    def export_to_csv(self, output_dir: str = "results") -> str:
        """Dump trades and performance to CSV."""
        os.makedirs(output_dir, exist_ok=True)
        trades_df = self.get_trade_summary()
        timestamp = format_timestamp(datetime.now())
        trades_file = os.path.join(output_dir, f"trades_{timestamp}.csv")

        # Create additional info table with trading configuration
        additional_info_df = self._create_additional_info_table()

        # Write both tables to the same CSV file
        with open(trades_file, 'w', newline='') as f:
            additional_info_df.to_csv(f, index=False)
            f.write('\n')  # Add empty line between tables
            trades_df.to_csv(f, index=False)

        return trades_file

    def create_optimized_excel_report(self, output_dir: str = "results") -> str:
        """Create FIXED Excel report with working trade data and larger text."""
        if not OPENPYXL_AVAILABLE:
            print("Warning: openpyxl not available. Please install it with: pip install openpyxl")
            return self.export_to_csv(output_dir)
            
        os.makedirs(output_dir, exist_ok=True)

        # Initialize components with larger layout
        wb = Workbook()
        ws = wb.active
        ws.title = "Backtest Results"

        style_manager = StyleManager(scale_factor=1.0)  # Already increased base font sizes
        layout_manager = LayoutManager(ws, max_columns=15)  # More columns for better layout
        table_builder = TableBuilder(layout_manager, style_manager)
        data_prep = DataPreparator(self)

        # Build report sections
        table_builder.create_title_section("BACKTEST RESULTS DASHBOARD")

        # Key metric highlight
        metrics = self.calculate_metrics()
        net_pnl_positive = metrics.net_pnl > 0 if metrics.net_pnl != 0 else None
        table_builder.create_highlight_metric(
            "TOTAL NET P&L", 
            f"₹{metrics.net_pnl:,.2f}",
            net_pnl_positive
        )

        # Summary metrics table
        metrics_data = data_prep.get_metrics_data()
        table_builder.create_metrics_table(metrics_data)

        # Configuration table
        config_data = self._create_additional_info_table()
        table_builder.create_config_table(config_data)

        # WORKING trades table with ALL columns and actual data
        trades_data = self.get_trade_summary()
        if not trades_data.empty:
            # Remove starting capital row for main display
            main_trades = trades_data.iloc[1:] if len(trades_data) > 1 else trades_data
            table_builder.create_trades_table(main_trades)

        # Save file
        timestamp = format_timestamp(datetime.now())
        filename = os.path.join(output_dir, f"Fixed_Backtest_Results_{timestamp}.xlsx")
        wb.save(filename)

        return filename

    # =====================================================
    # LEGACY METHODS (PRESERVED FOR COMPATIBILITY)
    # =====================================================

    def create_enhanced_excel_report(self, output_dir: str = "results") -> str:
        """Legacy method - redirects to fixed version."""
        return self.create_optimized_excel_report(output_dir)

    def export_to_excel(self, output_dir: str = "results") -> str:
        """Export to Excel - redirects to fixed version."""
        return self.create_optimized_excel_report(output_dir)

# Backward compatibility
BacktestResults = Results